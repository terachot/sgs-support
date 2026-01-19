import flet as ft
import asyncio
import os
import sys
import webbrowser
from openpyxl import load_workbook
from get_data_from_web import StudentScraper

# กำหนดค่าต่าง ๆ
students = []
#login_workpage = "https://sgs.bopp-obec.info/sgs/" #หน้าเข้าระบบ
#home_workpage = "https://sgs.bopp-obec.info/sgs/TblSchoolInfo/Show-TblSchoolInfo.aspx" #หน้าการใช้งานแรก
#all_workpage = "https://sgs.bopp-obec.info/sgs/TblTranscripts/Edit-TblTranscripts-Table.aspx" #หน้ากรอกตลอดภาคเรียน
#mid_workpage = "https://sgs.bopp-obec.info/sgs/TblTranscripts/Edit-TblTranscripts1-Table.aspx" #หน้ากรอกก่อนกลางภาค
#final_workpage = "https://sgs.bopp-obec.info/sgs/TblTranscripts/Edit-TblTranscripts2-Table.aspx" #หน้ากรอกหลังกลางภาค
#attribute_workpage = "https://sgs.bopp-obec.info/sgs/TblTranscriptsQ/Edit-TblTranscriptsQ-Table.aspx" #หน้ากรอกคุณลักษณะอันพึงประสงค์
#study_workpage = "https://sgs.bopp-obec.info/sgs/TblTranscriptsL/Edit-TblTranscriptsL-Table.aspx" #หน้ากรอกการอ่าน คิดวิเคราะห์ และเขียน

input_score = "input[id='ctl00_PageContent_TblTranscriptsSaveButton']" #บรรทึกข้อมูล คะแนน
#input_attribute = "input[id='ctl00_PageContent_TblTranscriptsQSaveButton']" #บรรทึกข้อมูล คุณลักษณะอันพึงประสงค์
#intput_study = "input[id='ctl00_PageContent_TblTranscriptsLSaveButton']" #บรรทึกข้อมูล การอ่าน คิดวิเคราะห์ และเขียน

# ค่าทดสอบ
login_workpage = "http://127.0.0.1:5500/" #หน้าเข้าระบบ
home_workpage = "http://127.0.0.1:5500/Show-TblSchoolInfo.html" #หน้าการใช้งานแรก
all_workpage = "http://127.0.0.1:5500/Edit-TblTranscripts-Table.html" #หน้ากรอกตลอดภาคเรียน
mid_workpage = "http://127.0.0.1:5500/Edit-TblTranscripts1-Table.html" #หน้ากรอกก่อนกลางภาค
final_workpage = "http://127.0.0.1:5500/Edit-TblTranscripts2-Table.html" #หน้ากรอกหลังกลางภาค
attribute_workpage = "http://127.0.0.1:5500/Edit-TblTranscriptsQ-Table.html" #หน้ากรอกคุณลักษณะอันพึงประสงค์
study_workpage = "http://127.0.0.1:5500/Edit-TblTranscriptsL-Table.html" #หน้ากรอกการอ่าน คิดวิเคราะห์ และเขียน

async def main(page: ft.Page):
    page.theme_mode = ft.ThemeMode.SYSTEM
    page.title = "SGS Support"
    page.window.height = 720
    page.window.width = 1280
    page.window.min_height = 540
    page.window.min_width = 960
    page.padding = ft.padding.symmetric(vertical=5, horizontal=20)
    page.window.center()
    page.vertical_alignment = ft.MainAxisAlignment.CENTER
    page.horizontal_alignment = ft.CrossAxisAlignment.CENTER

    scraper = StudentScraper(url=login_workpage)

    username = ft.TextField(label="ชื่อผู้ใช้",width=300, text_align=ft.TextAlign.CENTER)
    password = ft.TextField(label="รหัสผ่าน",width=300, text_align=ft.TextAlign.CENTER, password=True, can_reveal_password=True)
    result = ft.TextField(value="รอการเข้าสู่ระบบ", text_size=14, text_align=ft.TextAlign.LEFT, multiline=True, min_lines=2, max_lines=8 , expand=True)

    def resource_path(relative_path):
        """คืน path ที่ใช้ได้ทั้งตอนรันปกติและตอนรันจาก .exe"""
        if hasattr(sys, '_MEIPASS'):
            return os.path.join(sys._MEIPASS, relative_path)
        return os.path.join(os.path.abspath("."), relative_path)

    # ไฟล์ตัวอย่าง
    def open_file(e):
        #file_path = os.path.abspath("assets/sgs.xlsx")
        file_path = resource_path("assets/sgs.xlsx")
        webbrowser.open(f"file:///{file_path}")

    # ไฟล์จากผู้ใช้งาน
    file_picker = ft.FilePicker(on_result=lambda e: handle_file(e))
    page.overlay.append(file_picker)
    
    def handle_file(e: ft.FilePickerResultEvent):
        global students
        students.clear()
        if e.files:
            file_path = e.files[0].path  # ได้ path ของไฟล์ที่เลือก

            try:
                wb = load_workbook(file_path, read_only=True)
                sheet_names = wb.sheetnames

                for sheet_index, sheet_name in enumerate(sheet_names):
                    sheet = wb[sheet_name]
                    for i, row in enumerate(sheet.iter_rows(values_only=True)):
                        if row[0] is not None:
                            if sheet_index == 0 and i == 0:
                                students.append(list(row))  # ดึงหัวตารางจากชีตแรกเท่านั้น
                            elif i > 0:
                                students.append(list(row))  # ข้ามหัวตารางในชีตอื่น

                students_header = students[0]
                students.pop(0)
                students = sorted(students, key=lambda x: int(x[0]))
                students.insert(0, students_header)
            except Exception as ex:
                result.value = f"เกิดข้อผิดพลาด: {ex}"
            result.value = f"โหลดข้อมูลนักเรียนทั้งหมด {len(students)-1} คน"
            page.update()

    async def theme_changed(e):
        page.theme_mode = (
            ft.ThemeMode.DARK if page.theme_mode == ft.ThemeMode.LIGHT else ft.ThemeMode.LIGHT
        )
        theme_button.label = (
            "Light theme" if page.theme_mode == ft.ThemeMode.LIGHT else "Dark theme"
        )
        page.update()
    theme_button = ft.Switch(label="Light theme", on_change=theme_changed)

    async def goto_webpage(e,wp):
        await scraper.change_page(wp)

    def make_button(url):
        async def handler(e):
            await goto_webpage(e, url)
        return ft.ElevatedButton(text="เปิดหน้าเว็บ", on_click=handler,width=300,height=45, bgcolor=ft.Colors.BLUE_400, color=ft.Colors.GREY_900)

    async def work_page(e):
        login_row.visible = False
        tab_row.visible = True
        logout.visible = True
        page.update()
    
    async def login_page(e):
        tab_row.visible = False
        login_row.visible = True
        logout.visible = False
        page.update()

    async def logout_menu(e):
        await scraper.logout_web()
        await login_page(e)
        await close_web(e)
        result.value = "ออกจากระบบสำเร็จ"
        page.update()
    logout = ft.ElevatedButton(text="logout", on_click=logout_menu)
    logout.visible = False

    async def open_web(e):
        if scraper.browser is None:
            await scraper.open_web()
            await scraper.login_web(username.value, password.value)
            wt = 0
            while await scraper.get_url() != home_workpage:
                wt += 1
                await asyncio.sleep(0.6)
                if wt == 30:
                    result.value = "การเชื่อมต่อใช้เวลานานเกินไป"
                    page.update()
                    raise Exception("การเชื่อมต่อใช้เวลานานเกินไป")

            if await scraper.get_url() == home_workpage:
                await work_page(e)
                result.value = "เข้าใช้งานสำเร็จ"
                page.update()
            else:
                await close_web(e)
                result.value = f"ผิดพลาด เข้าใช้งานไม่สำเร็จ"
                page.update()
        else:
            #print("Browser เปิดอยู่แล้ว")
            result.value = "Browser เปิดอยู่แล้ว"

    async def close_web(e):
        if not scraper.page :
            scraper.close_web()
            result.value += f"\nปิดเว็บเรียบร้อยแล้ว"
            page.update()
        else:
            result.value += f"\nยังไม่มี browser ที่เปิดอยู่"
            page.update()
    
    #async def windows_close(e):
    #    print("Shutting down...")
    #    await asyncio.sleep(0.1)

    async def score_before(e):
        if await scraper.get_url() == mid_workpage:
            try:
                result.value = ""
                web_datas = await scraper.get_data() #ข้อมูลทั้งหมดจากหน้าเว็บมี index, id, name
                filled_count = 0 #จำนวนคนที่กรอกสำเร็จ
                web_index = 0 #แถวที่กำลังทำการกรอกในเว็บ
                excel_index = 0 #แถวข้อมูลใน excel
                headers = students[0] #กำหนดหัวตาราง
                i = len(students) - 1
                #print(f"จำนวน นร. : {i} type{type(i)}")
                while web_index < i and excel_index <= i:
                    # ดึงค่ารหัสนักเรียนจากในเว็บด้วย web_index จาก web_datas
                    w_id = int(web_datas[web_index]["id"])
                    # ดึงค่ารหัสนักเรียนจากตัวแปร students ด้วยตัวแปร excel_index
                    student = students[excel_index + 1]
                    e_id = student[0]
                    #print(f"web id : {w_id} web index : {web_index}\nexcel id : {e_id} excel index : {web_index}")
                    # เทียบค่ารหัสนักเรียนจาก w_id กับ e_id
                    if w_id == e_id: # ถ้าตรงกันให้กรอกข้อมูลในแถวนั้น แล้วเพิ่มค่า web_index และ excel_index ไป 1 ค่า
                        for j, key in enumerate(headers):
                            if key in ["stdID", "student Name","S10","S11","S12","S13","S14","S15","S16","S17","S18","Final","Q1","Q2","Q3","Q4","Q5","Q6","Q7","Q8","L1","L2","L3","L4","L5"]:
                                continue  # ข้ามช่องที่ไม่ใช่ input
                            value = student[j]
                            if value is not None:
                                input_id = f"ctl00_PageContent_TblTranscriptsTableControlRepeater_ctl{str(web_index+1).zfill(2)}_{key}"
                                await scraper.page.fill(f"#{input_id}", str(value))
                        filled_count += 1
                        web_index += 1 # เลื่อนแถวตารางกรอกถัดไป
                        excel_index += 1 # เลื่อน id ในไฟล์ถัดไป
                    # ถ้าไม่ตรงให้เพิ่มค่าตัวแปร excel_index ไป 1 ค่า
                    elif w_id < e_id: # แต่ถ้าค่า id ที่จะกรอกมากกว่า id ในเว็บ
                        web_index += 1 # เลื่อนแถวตารางกรอกถัดไป
                    else: # ถ้าค่า id ที่จะกรอกน้อยกว่า id ในเว็บให้ข้ามเป็น id ถัดไป
                        excel_index += 1 # เลื่อน id ในไฟล์ถัดไป
                
                await scraper.page.locator(input_score).click()
                result.value += f"\nลงข้อมูลก่อนกลางภาคเสร็จสิ้น\nลงข้อมูลเสร็จสิ้นทั้งหมด {filled_count} คน"
                page.update()
            except Exception as ex:
                result.value += f"\nลงเข้ามูลก่อนกลางภาคไม่สำเร็จ : {ex}"
                page.update()
        else:
            #print("ไม่ใช่หน้าที่จะทำการกรอกคะแนน")
            result.value = "ไม่ใช่หน้าที่จะทำการกรอกคะแนน"
            page.update()

    async def score_after(e):
        if await scraper.get_url() == final_workpage:
            try:
                result.value = ""
                web_datas = await scraper.get_data() #ข้อมูลทั้งหมดจากหน้าเว็บมี index, id, name
                filled_count = 0 #จำนวนคนที่กรอกสำเร็จ
                web_index = 0 #แถวที่กำลังทำการกรอกในเว็บ
                excel_index = 0 #แถวข้อมูลใน excel
                headers = students[0] #กำหนดหัวตาราง
                i = len(students) - 1
                #print(f"จำนวน นร. : {i} type{type(i)}")
                while web_index < i and excel_index <= i:
                    # ดึงค่ารหัสนักเรียนจากในเว็บด้วย web_index จาก web_datas
                    w_id = int(web_datas[web_index]["id"])
                    # ดึงค่ารหัสนักเรียนจากตัวแปร students ด้วยตัวแปร excel_index
                    student = students[excel_index + 1]
                    e_id = student[0]
                    #print(f"web id : {w_id} web index : {web_index}\nexcel id : {e_id} excel index : {web_index}")
                    # เทียบค่ารหัสนักเรียนจาก w_id กับ e_id
                    if w_id == e_id: # ถ้าตรงกันให้กรอกข้อมูลในแถวนั้น แล้วเพิ่มค่า web_index และ excel_index ไป 1 ค่า
                        for j, key in enumerate(headers):
                            if key in ["stdID", "student Name","S1","S2","S3","S4","S5","S6","S7","S8","S9","Midterm","Q1","Q2","Q3","Q4","Q5","Q6","Q7","Q8","L1","L2","L3","L4","L5"]:
                                continue  # ข้ามช่องที่ไม่ใช่ input
                            value = student[j]
                            if value is not None:
                                input_id = f"ctl00_PageContent_TblTranscriptsTableControlRepeater_ctl{str(web_index+1).zfill(2)}_{key}"
                                await scraper.page.fill(f"#{input_id}", str(value))
                        filled_count += 1
                        web_index += 1 # เลื่อนแถวตารางกรอกถัดไป
                        excel_index += 1 # เลื่อน id ในไฟล์ถัดไป
                    # ถ้าไม่ตรงให้เพิ่มค่าตัวแปร excel_index ไป 1 ค่า
                    elif w_id < e_id: # แต่ถ้าค่า id ที่จะกรอกมากกว่า id ในเว็บ
                        web_index += 1 # เลื่อนแถวตารางกรอกถัดไป
                    else: # ถ้าค่า id ที่จะกรอกน้อยกว่า id ในเว็บให้ข้ามเป็น id ถัดไป
                        excel_index += 1 # เลื่อน id ในไฟล์ถัดไป
                
                await scraper.page.locator(input_score).click()
                result.value += f"\nลงข้อมูลก่อนกลางภาคเสร็จสิ้น\nลงข้อมูลเสร็จสิ้นทั้งหมด {filled_count} คน"
                page.update()
            except Exception as ex:
                result.value += f"\nลงเข้ามูลก่อนกลางภาคไม่สำเร็จ : {ex}"
                page.update()
        else:
            #print("ไม่ใช่หน้าที่จะทำการกรอกคะแนน")
            result.value = "ไม่ใช่หน้าที่จะทำการกรอกคะแนน"
            page.update()

    async def score_all(e):
        if await scraper.get_url() == all_workpage:
            try:
                result.value = ""
                web_datas = await scraper.get_data() #ข้อมูลทั้งหมดจากหน้าเว็บมี index, id, name
                filled_count = 0 #จำนวนคนที่กรอกสำเร็จ
                web_index = 0 #แถวที่กำลังทำการกรอกในเว็บ
                excel_index = 0 #แถวข้อมูลใน excel
                headers = students[0] #กำหนดหัวตาราง
                i = len(students) - 1
                #print(f"จำนวน นร. : {i} type{type(i)}")
                while web_index < i and excel_index <= i:
                    # ดึงค่ารหัสนักเรียนจากในเว็บด้วย web_index จาก web_datas
                    w_id = int(web_datas[web_index]["id"])
                    # ดึงค่ารหัสนักเรียนจากตัวแปร students ด้วยตัวแปร excel_index
                    student = students[excel_index + 1]
                    e_id = student[0]
                    #print(f"web id : {w_id} web index : {web_index}\nexcel id : {e_id} excel index : {web_index}")
                    # เทียบค่ารหัสนักเรียนจาก w_id กับ e_id
                    if w_id == e_id: # ถ้าตรงกันให้กรอกข้อมูลในแถวนั้น แล้วเพิ่มค่า web_index และ excel_index ไป 1 ค่า
                        for j, key in enumerate(headers):
                            if key in ["stdID", "student Name","S4","S5","S6","S7","S8","S9","S13","S14","S15","S16","S17","S18","Q1","Q2","Q3","Q4","Q5","Q6","Q7","Q8","L1","L2","L3","L4","L5"]:
                                continue  # ข้ามช่องที่ไม่ใช่ input
                            value = student[j]
                            if value is not None: # ยังไม่ได้ตรวจสอบ id ก่อนลงข้อมูล
                                input_id = f"ctl00_PageContent_TblTranscriptsTableControlRepeater_ctl{str(web_index+1).zfill(2)}_{key}"
                                await scraper.page.fill(f"#{input_id}", str(value))
                        filled_count += 1
                        web_index += 1 # เลื่อนแถวตารางกรอกถัดไป
                        excel_index += 1 # เลื่อน id ในไฟล์ถัดไป
                    # ถ้าไม่ตรงให้เพิ่มค่าตัวแปร excel_index ไป 1 ค่า
                    elif w_id < e_id: # แต่ถ้าค่า id ที่จะกรอกมากกว่า id ในเว็บ
                        web_index += 1 # เลื่อนแถวตารางกรอกถัดไป
                    else: # ถ้าค่า id ที่จะกรอกน้อยกว่า id ในเว็บให้ข้ามเป็น id ถัดไป
                        excel_index += 1 # เลื่อน id ในไฟล์ถัดไป
                
                await scraper.page.locator(input_score).click()
                result.value += f"\nลงข้อมูลก่อนกลางภาคเสร็จสิ้น\nลงข้อมูลเสร็จสิ้นทั้งหมด {filled_count} คน"
                page.update()
            except Exception as ex:
                result.value += f"\nลงเข้ามูลก่อนกลางภาคไม่สำเร็จ : {ex}"
                page.update()
        else:
            #print("ไม่ใช่หน้าที่จะทำการกรอกคะแนน")
            result.value = "ไม่ใช่หน้าที่จะทำการกรอกคะแนน"
            page.update()

    async def score_attribute(e):
        if await scraper.get_url() == attribute_workpage:
            try:
                result.value = ""
                web_datas = await scraper.get_data() #ข้อมูลทั้งหมดจากหน้าเว็บมี index, id, name
                filled_count = 0 #จำนวนคนที่กรอกสำเร็จ
                web_index = 0 #แถวที่กำลังทำการกรอกในเว็บ
                excel_index = 0 #แถวข้อมูลใน excel
                headers = students[0] #กำหนดหัวตาราง
                i = len(students) - 1
                #print(f"จำนวน นร. : {i} type{type(i)}")
                while web_index < i and excel_index <= i:
                    # ดึงค่ารหัสนักเรียนจากในเว็บด้วย web_index จาก web_datas
                    w_id = int(web_datas[web_index]["id"])
                    # ดึงค่ารหัสนักเรียนจากตัวแปร students ด้วยตัวแปร excel_index
                    student = students[excel_index + 1]
                    e_id = student[0]
                    #print(f"web id : {w_id} web index : {web_index}\nexcel id : {e_id} excel index : {web_index}")
                    # เทียบค่ารหัสนักเรียนจาก w_id กับ e_id
                    if w_id == e_id: # ถ้าตรงกันให้กรอกข้อมูลในแถวนั้น แล้วเพิ่มค่า web_index และ excel_index ไป 1 ค่า
                        for j, key in enumerate(headers):
                            if key in ["stdID", "student Name","S1","S2","S3","S4","S5","S6","S7","S8","S9","Midterm","S10","S11","S12","S13","S14","S15","S16","S17","S18","Final","L1","L2","L3","L4","L5"]:
                                continue  # ข้ามช่องที่ไม่ใช่ input
                            value = student[j]
                            if value is not None: # ยังไม่ได้ตรวจสอบ id ก่อนลงข้อมูล
                                input_id = f"ctl00_PageContent_TblTranscriptsQTableControlRepeater_ctl{str(web_index+1).zfill(2)}_{key}"
                                await scraper.page.fill(f"#{input_id}", str(value))
                        filled_count += 1
                        web_index += 1 # เลื่อนแถวตารางกรอกถัดไป
                        excel_index += 1 # เลื่อน id ในไฟล์ถัดไป
                    # ถ้าไม่ตรงให้เพิ่มค่าตัวแปร excel_index ไป 1 ค่า
                    elif w_id < e_id: # แต่ถ้าค่า id ที่จะกรอกมากกว่า id ในเว็บ
                        web_index += 1 # เลื่อนแถวตารางกรอกถัดไป
                    else: # ถ้าค่า id ที่จะกรอกน้อยกว่า id ในเว็บให้ข้ามเป็น id ถัดไป
                        excel_index += 1 # เลื่อน id ในไฟล์ถัดไป
                
                await scraper.page.locator(input_score).click()
                result.value += f"\nลงข้อมูลก่อนกลางภาคเสร็จสิ้น\nลงข้อมูลเสร็จสิ้นทั้งหมด {filled_count} คน"
                page.update()
            except Exception as ex:
                result.value += f"\nลงเข้ามูลก่อนกลางภาคไม่สำเร็จ : {ex}"
                page.update()
        else:
            #print("ไม่ใช่หน้าที่จะทำการกรอกคะแนน")
            result.value = "ไม่ใช่หน้าที่จะทำการกรอกคะแนน"
            page.update()
    
    async def score_study(e):
        if await scraper.get_url() == study_workpage:
            try:
                result.value = ""
                web_datas = await scraper.get_data() #ข้อมูลทั้งหมดจากหน้าเว็บมี index, id, name
                filled_count = 0 #จำนวนคนที่กรอกสำเร็จ
                web_index = 0 #แถวที่กำลังทำการกรอกในเว็บ
                excel_index = 0 #แถวข้อมูลใน excel
                headers = students[0] #กำหนดหัวตาราง
                i = len(students) - 1
                #print(f"จำนวน นร. : {i} type{type(i)}")
                while web_index < i and excel_index <= i:
                    # ดึงค่ารหัสนักเรียนจากในเว็บด้วย web_index จาก web_datas
                    w_id = int(web_datas[web_index]["id"])
                    # ดึงค่ารหัสนักเรียนจากตัวแปร students ด้วยตัวแปร excel_index
                    student = students[excel_index + 1]
                    e_id = student[0]
                    #print(f"web id : {w_id} web index : {web_index}\nexcel id : {e_id} excel index : {web_index}")
                    # เทียบค่ารหัสนักเรียนจาก w_id กับ e_id
                    if w_id == e_id: # ถ้าตรงกันให้กรอกข้อมูลในแถวนั้น แล้วเพิ่มค่า web_index และ excel_index ไป 1 ค่า
                        for j, key in enumerate(headers):
                            if key in ["stdID", "student Name","S1","S2","S3","S4","S5","S6","S7","S8","S9","Midterm","S10","S11","S12","S13","S14","S15","S16","S17","S18","Final","Q1","Q2","Q3","Q4","Q5","Q6","Q7","Q8"]:
                                continue  # ข้ามช่องที่ไม่ใช่ input
                            value = student[j]
                            if value is not None: # ยังไม่ได้ตรวจสอบ id ก่อนลงข้อมูล
                                input_id = f"ctl00_PageContent_TblTranscriptsLTableControlRepeater_ctl{str(web_index+1).zfill(2)}_{key}"
                                await scraper.page.fill(f"#{input_id}", str(value))
                        filled_count += 1
                        web_index += 1 # เลื่อนแถวตารางกรอกถัดไป
                        excel_index += 1 # เลื่อน id ในไฟล์ถัดไป
                    # ถ้าไม่ตรงให้เพิ่มค่าตัวแปร excel_index ไป 1 ค่า
                    elif w_id < e_id: # แต่ถ้าค่า id ที่จะกรอกมากกว่า id ในเว็บ
                        web_index += 1 # เลื่อนแถวตารางกรอกถัดไป
                    else: # ถ้าค่า id ที่จะกรอกน้อยกว่า id ในเว็บให้ข้ามเป็น id ถัดไป
                        excel_index += 1 # เลื่อน id ในไฟล์ถัดไป
                
                await scraper.page.locator(input_score).click()
                result.value += f"\nลงข้อมูลก่อนกลางภาคเสร็จสิ้น\nลงข้อมูลเสร็จสิ้นทั้งหมด {filled_count} คน"
                page.update()
            except Exception as ex:
                result.value += f"\nลงเข้ามูลก่อนกลางภาคไม่สำเร็จ : {ex}"
                page.update()
        else:
            #print("ไม่ใช่หน้าที่จะทำการกรอกคะแนน")
            result.value = "ไม่ใช่หน้าที่จะทำการกรอกคะแนน"
            page.update()

    header_body = [
        logout,
        ft.Text(value="SGS Support", size=20),
        #ft.ElevatedButton(text=" work ", on_click=work_page),
        #ft.ElevatedButton(text="login", on_click=login_page),
        #ft.ElevatedButton(text="ปิดเว็บ", on_click=close_web),
        #ft.ElevatedButton(text="เลือกไฟล์ Excel",on_click=lambda _: file_picker.pick_files(allow_multiple=False, allowed_extensions=["xlsx"])),
        theme_button,
    ]
    
    login_body = [
        ft.Text(value="ลงชื่อ", size=18, text_align=ft.TextAlign.CENTER),
        username,
        password,
        ft.ElevatedButton(text="เข้าใช้งาน",on_click=open_web,width=300,height=50, bgcolor=ft.Colors.ORANGE_300, color=ft.Colors.GREY_900),
    ]

    footer_body = [
        ft.Text(value="การทำงาน", text_align=ft.TextAlign.LEFT,),
        result,
    ]

    header_row = ft.Row(
        header_body,
        alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
        expand=1,
        spacing=5,
    )
    login_row = ft.Column(
        login_body,
        horizontal_alignment=ft.CrossAxisAlignment.CENTER,
        alignment=ft.MainAxisAlignment.CENTER,
        expand=5,
        spacing=15,
    )

    footer_row = ft.Column(
        footer_body,
        expand=2,
        spacing=5,
    )

    tab_row = ft.Tabs(
        selected_index=0,
        animation_duration=300,
        tabs=[
            ft.Tab(
                text="         การใช้งาน         ",
                icon=ft.Icons.HOME,
                content=ft.Container(
                    content=ft.Column(
                        [
                            ft.Row(
                                [
                                ft.ElevatedButton(text="เลือกไฟล์ Excel",on_click=lambda _: file_picker.pick_files(allow_multiple=False, allowed_extensions=["xlsx"]),width=300,height=40, bgcolor=ft.Colors.GREEN_300, color=ft.Colors.GREY_900),
                                ft.ElevatedButton(text="โหลดไฟล์ตัวอย่าง",on_click=open_file,width=300,height=40, bgcolor=ft.Colors.ORANGE_300, color=ft.Colors.GREY_900),
                                ],
                                alignment=ft.MainAxisAlignment.CENTER,
                            ),
                            ft.DataTable(
                                heading_row_height=0,
                                column_spacing=30,
                                data_row_min_height=40,
                                data_row_max_height=50,
                                columns=[
                                    ft.DataColumn(ft.Text("")),
                                    ft.DataColumn(ft.Text("")),
                                    ft.DataColumn(ft.Text("")),
                                    ft.DataColumn(ft.Text("")),
                                    ft.DataColumn(ft.Text("")),
                                    ft.DataColumn(ft.Text("")),
                                ],
                                rows=[
                                    ft.DataRow(
                                        cells=[
                                            ft.DataCell(ft.Text("stdID")),
                                            ft.DataCell(ft.Text("หมายถึง")),
                                            ft.DataCell(ft.Text("รหัสนักเรียน")),
                                            ft.DataCell(ft.Text("student Name")),
                                            ft.DataCell(ft.Text("หมายถึง")),
                                            ft.DataCell(ft.Text("ชื่อ-นามสกุล นักเรียน")),
                                        ],
                                    ),
                                    ft.DataRow(
                                        cells=[
                                            ft.DataCell(ft.Text("S1 - S9")),
                                            ft.DataCell(ft.Text("หมายถึง")),
                                            ft.DataCell(ft.Text("คะแนนก่อนกลางภาค")),
                                            ft.DataCell(ft.Text("S10 - S18")),
                                            ft.DataCell(ft.Text("หมายถึง")),
                                            ft.DataCell(ft.Text("คะแนนหลังกลางภาค")),
                                        ],
                                    ),
                                    ft.DataRow(
                                        cells=[
                                            ft.DataCell(ft.Text("Midterm")),
                                            ft.DataCell(ft.Text("หมายถึง")),
                                            ft.DataCell(ft.Text("คะแนนสอบกลางถาค")),
                                            ft.DataCell(ft.Text("Final")),
                                            ft.DataCell(ft.Text("หมายถึง")),
                                            ft.DataCell(ft.Text("คะแนนสอบปลายภาค")),
                                        ],
                                    ),
                                    ft.DataRow(
                                        cells=[
                                            ft.DataCell(ft.Text("Q1 - Q8")),
                                            ft.DataCell(ft.Text("หมายถึง")),
                                            ft.DataCell(ft.Text("คะแนนคุณลักษณะอันพึงประสงค์")),
                                            ft.DataCell(ft.Text("L1 - L5")),
                                            ft.DataCell(ft.Text("หมายถึง")),
                                            ft.DataCell(ft.Text("คะแนนการอ่าน คิดวิเคราะห์ และเขียน")),
                                        ],
                                    ),
                                ]
                            ),
                        ],
                        alignment=ft.MainAxisAlignment.CENTER,
                        horizontal_alignment=ft.CrossAxisAlignment.CENTER,
                        spacing=15,
                        scroll=ft.ScrollMode.ALWAYS,
                    ),
                    alignment=ft.alignment.center,
                    padding=20,
                ),
            ),
            ft.Tab(
                text="    ลงคะแนน ก่อนกลางภาค    ",
                icon=ft.Icons.BACKUP_TABLE_OUTLINED,
                content=ft.Container(
                    content=ft.Column(
                        [
                            make_button(mid_workpage),
                            ft.ElevatedButton(text="ลงคะแนนก่อนกลางภาค",on_click=score_before,width=300,height=45, bgcolor=ft.Colors.ORANGE_300, color=ft.Colors.GREY_900),
                        ],
                        alignment=ft.MainAxisAlignment.CENTER,
                        horizontal_alignment=ft.CrossAxisAlignment.CENTER,
                        spacing=15,
                    ),
                    alignment=ft.alignment.center,
                    padding=20,
                ),
            ),
            ft.Tab(
                text="   ลงคะแนน หลังกลางภาค   ",
                icon=ft.Icons.BACKUP_TABLE_ROUNDED,
                content=ft.Container(
                    content=ft.Column(
                        [
                            make_button(final_workpage),
                            ft.ElevatedButton(text="ลงคะแนนหลังกลางภาค",on_click=score_after,width=300,height=45, bgcolor=ft.Colors.ORANGE_300, color=ft.Colors.GREY_900),
                        ],
                        alignment=ft.MainAxisAlignment.CENTER,
                        horizontal_alignment=ft.CrossAxisAlignment.CENTER,
                        spacing=15,
                    ),
                    alignment=ft.alignment.center,
                    padding=20,
                ),
            ),
            ft.Tab(
                text="   ลงคะแนน ตลอดภาคเรียน   ",
                icon=ft.Icons.BACKUP_TABLE_ROUNDED,
                content=ft.Container(
                    content=ft.Column(
                        [
                            make_button(all_workpage),
                            ft.ElevatedButton(text="ลงคะแนนตลอดภาคเรียน",on_click=score_all,width=300,height=45, bgcolor=ft.Colors.RED_300, color=ft.Colors.GREY_900),
                        ],
                        alignment=ft.MainAxisAlignment.CENTER,
                        horizontal_alignment=ft.CrossAxisAlignment.CENTER,
                        spacing=15,
                    ),
                    alignment=ft.alignment.center,
                    padding=20,
                ),
            ),
            ft.Tab(
                text="   คุณลักษณะอันพึงประสงค์   ",
                icon=ft.Icons.SPORTS_MARTIAL_ARTS_OUTLINED,
                content=ft.Container(
                    content=ft.Column(
                        [
                            make_button(attribute_workpage),
                            ft.ElevatedButton(text="ลงคะแนนคุณลักษณะอันพึงประสงค์",on_click=score_attribute,width=300,height=45, bgcolor=ft.Colors.ORANGE_300, color=ft.Colors.GREY_900),
                        ],
                        alignment=ft.MainAxisAlignment.CENTER,
                        horizontal_alignment=ft.CrossAxisAlignment.CENTER,
                        spacing=15,
                    ),
                    alignment=ft.alignment.center,
                    padding=20,
                ),
            ),
            ft.Tab(
                text="การอ่าน คิดวิเคราะห์ และเขียน",
                icon=ft.Icons.CHROME_READER_MODE_OUTLINED,
                content=ft.Container(
                    content=ft.Column(
                        [
                            #ft.Text("การอ่าน คิดวิเคราะห์ และเขียน", size=20),
                            make_button(study_workpage),
                            ft.ElevatedButton(text="ลงคะแนนการอ่าน คิดวิเคราะห์ และเขียน",on_click=score_study,width=300,height=45, bgcolor=ft.Colors.ORANGE_300, color=ft.Colors.GREY_900),
                        ],
                        alignment=ft.MainAxisAlignment.CENTER,
                        horizontal_alignment=ft.CrossAxisAlignment.CENTER,
                        spacing=15,
                    ),
                    alignment=ft.alignment.center,
                    padding=20,
                ),
            ),
        ],
        expand=6,
    )
    tab_row.visible = False
    
    #page.on_close = windows_close

    page.add(
        header_row,
        login_row,
        tab_row,
        footer_row,
    )

asyncio.run(ft.app(main))